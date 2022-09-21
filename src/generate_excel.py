#!/usr/bin/env python
# -*- coding: utf-8 -*-
# generate formatted excel file from internal csv file
# csv_to_xlsx(csv_file, excel_file, sheet_name, anchor='ANCHOR', test='TEST', type = 'psnr'):
# inputs: csv_file, excel_file , sheet_name, anchor/test name and calc type('psnr' and 'all' are supported now)
# outputs: formatted excel file
# requirments: openpyxl libs must be installed before called this function
# ==============================================================================================================
import sys
import os
import csv
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

def my_border(t_border, b_border, l_border, r_border):
    border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                    bottom=Side(border_style=b_border, color=colors.BLACK),
                    left=Side(border_style=l_border, color=colors.BLACK),
                    right=Side(border_style=r_border, color=colors.BLACK))
    return border

def format_border(sheet, s_column, s_index, e_column , e_index):
    for row in tuple(sheet[s_column + str(s_index):e_column + str(e_index)]):
        for cell in row:
            cell.border = my_border('none', 'none', 'none', 'none')

def set_solid_border(sheet, area_list):
    for area in area_list:
        s_column = area[0]
        s_index = area[1]
        e_column = area[2]
        e_index = area[3]

        for cell in sheet[s_column][s_index - 1:e_index]:
            cell.border = my_border(cell.border.top.style, cell.border.bottom.style,
                                    'medium', cell.border.right.style)

        for cell in sheet[e_column][s_index - 1:e_index]:
            cell.border = my_border(cell.border.top.style, cell.border.bottom.style,
                                    cell.border.left.style, 'medium')

        for row in tuple(sheet[s_column + str(s_index):e_column + str(s_index)]):
            for cell in row:
                cell.border = my_border('medium', cell.border.bottom.style,
                                        cell.border.left.style, cell.border.right.style)

        for row in tuple(sheet[s_column + str(e_index):e_column + str(e_index)]):
            for cell in row:
                cell.border = my_border(cell.border.top.style, 'medium',
                                        cell.border.left.style, cell.border.right.style)
def get_delta_endlabel(type):
    if type == "all":
        delta = 12
        end_label = 'AI'
    elif type == 'psnr':
        delta = 7
        end_label = 'T'
    return end_label, delta

def gen_lable_index(type='psnr'):
    lable_index = dict()
    if type == 'all':
        lable_index['anchor'] = ['C1:M1','C','M',3]
        lable_index['test'] = ['O1:Y1','O','Y',15]
        lable_index['BD-PSNR'] = ['AA1:AD1','AA','AD',27]
        lable_index['BD-SSIM'] = ['AE1:AH1','AE','AH',31]
        lable_index['BD-VMAF'] = ['AI1:AI1','AI','AI',35]
        lable_index['SPEED-UP'] = ['AJ1:AJ1','AJ','AJ',36]
    elif type == 'psnr':
        lable_index['anchor'] = ['C1:H1','C','H',3]
        lable_index['test'] = ['J1:O1','J','O',10]
        lable_index['BD-PSNR'] = ['Q1:T1','Q','T',17]
        lable_index['SPEED-UP'] = ['U1:U1','U','U',21]
    else:
        print("only all and psnr are supported now!")
    return lable_index

def draw_header(sheet, label, anchor, test, type='psnr'):
    font = Font(name='Times New Roman', bold = True)
    align = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(label['anchor'][0])
    sheet.cell(row=1, column=label['anchor'][3]).value = anchor
    sheet.cell(row=1, column=label['anchor'][3]).font = font
    sheet.cell(row=1, column=label['anchor'][3]).alignment = align
    sheet.merge_cells(label['test'][0])
    sheet.cell(row=1, column=label['test'][3]).value = test
    sheet.cell(row=1, column=label['test'][3]).font = font
    sheet.cell(row=1, column=label['test'][3]).alignment = align
    sheet.merge_cells(label['BD-PSNR'][0])
    sheet.cell(row=1, column=label['BD-PSNR'][3]).value = 'BD-PSNR'
    sheet.cell(row=1, column=label['BD-PSNR'][3]).font = font
    sheet.cell(row=1, column=label['BD-PSNR'][3]).alignment = align
    sheet.column_dimensions[label['SPEED-UP'][1]].width = 15.0
    sheet.cell(row=1,column=label['SPEED-UP'][3]).value = 'SPEED-UP'
    sheet.cell(row=1,column=label['SPEED-UP'][3]).font = font
    sheet.cell(row=1,column=label['SPEED-UP'][3]).alignment = align
    if type == 'all':
        sheet.merge_cells(label['BD-SSIM'][0])
        sheet.cell(row=1, column=label['BD-SSIM'][3]).value = 'BD-SSIM'
        sheet.cell(row=1, column=label['BD-SSIM'][3]).font = font
        sheet.cell(row=1, column=label['BD-SSIM'][3]).alignment = align
        sheet.column_dimensions[label['BD-VMAF'][1]].width = 15.0
        sheet.cell(row=1,column=label['BD-VMAF'][3]).value = 'BD-VMAF'
        sheet.cell(row=1,column=label['BD-VMAF'][3]).font = font
        sheet.cell(row=1,column=label['BD-VMAF'][3]).alignment = align

def draw_second_row(sheet, label, anchor, test, type='psnr'):
    end_label, delta = get_delta_endlabel(type)
    sheet.cell(row=2, column=1).value = 'sequence'
    sheet.column_dimensions['B'].width = 50.0
    sheet.cell(row=2, column=2).value = sheet.cell(row=2, column=(2+delta)).value = 'kbps'
    sheet.cell(row=2, column=3).value = sheet.cell(row=2, column=(3+delta)).value = 'Y-PSNR'
    sheet.cell(row=2, column=4).value = sheet.cell(row=2, column=(4+delta)).value = 'U-PSNR'
    sheet.cell(row=2, column=5).value = sheet.cell(row=2, column=(5+delta)).value = 'V-PSNR'
    sheet.cell(row=2, column=6).value = sheet.cell(row=2, column=(6+delta)).value = 'C-PSNR'
    if type == 'all':
        sheet.cell(row=2, column=7).value = sheet.cell(row=2, column=(7+delta)).value = 'Y-SSIM'
        sheet.cell(row=2, column=8).value = sheet.cell(row=2, column=(8+delta)).value = 'U-SSIM'
        sheet.cell(row=2, column=9).value = sheet.cell(row=2, column=(9+delta)).value = 'V-SSIM'
        sheet.cell(row=2, column=10).value = sheet.cell(row=2, column=(10+delta)).value = 'C-SSIM'
        sheet.cell(row=2, column=11).value = sheet.cell(row=2, column=(11+delta)).value = 'VMAF'
        sheet.cell(row=2, column=12).value = sheet.cell(row=2, column=(12+delta)).value = 'EncFPS'
        sheet.cell(row=2, column=26).value = 'Y'
        sheet.cell(row=2, column=27).value = 'U'
        sheet.cell(row=2, column=28).value = 'V'
        sheet.cell(row=2, column=29).value = 'C'
        sheet.cell(row=2, column=30).value = 'Y'
        sheet.cell(row=2, column=31).value = 'U'
        sheet.cell(row=2, column=32).value = 'V'
        sheet.cell(row=2, column=33).value = 'C'
        sheet.cell(row=2, column=34).value = 'YUV'
        sheet.cell(row=2, column=35).value = 'RATIO'
    elif type == 'psnr':
        sheet.cell(row=2, column=7).value = sheet.cell(row=2, column=(7+delta)).value = 'EncFPS'
        sheet.cell(row=2, column=16).value = 'Y'
        sheet.cell(row=2, column=17).value = 'U'
        sheet.cell(row=2, column=18).value = 'V'
        sheet.cell(row=2, column=19).value = 'C'
        sheet.cell(row=2, column=20).value = 'RATIO'

def format_excel(sheet, label, data_rows, all_rows, type = 'psnr'):
    end_label, delta = get_delta_endlabel(type)
    align = Alignment(horizontal='center', vertical='center')
    font = Font(name='Arial', bold = False, size=9)
    for eachCommonRow in sheet["A2:"+ label['SPEED-UP'][2]+ str(all_rows)]:
        for eachCellInRow in eachCommonRow:
            eachCellInRow.font = font
            eachCellInRow.alignment = align

    format_border(sheet, 'A', 1, label['SPEED-UP'][1], all_rows)
    #first two lines
    set_solid_border(sheet,[[label['anchor'][1], 1, label['anchor'][2], 1], \
        [label['test'][1], 1, label['test'][2], 1], \
        [label['BD-PSNR'][1], 1, label['BD-PSNR'][2], 1], \
        [label['SPEED-UP'][1], 1, label['SPEED-UP'][2], 1]])
    set_solid_border(sheet,[['A', 2, 'A', 2], \
        [label['anchor'][1], 2, label['anchor'][2], 2], \
        [label['test'][1], 2, label['test'][2], 2], \
        [label['BD-PSNR'][1], 2, label['BD-PSNR'][2], 2], \
        [label['SPEED-UP'][1], 2, label['SPEED-UP'][2], 2]])
    
    for row in range(3, data_rows+1, 4):
        set_solid_border(sheet,[['A', row, label['anchor'][2], row+3], \
            [label['test'][1], row, label['test'][2], row+3], \
            [label['BD-PSNR'][1], row, label['SPEED-UP'][2], row+3]])
    #vertical border
    set_solid_border(sheet,[['A', 3, 'A', data_rows], \
        ['B', 2, 'B', data_rows], ['C', 2, 'F', data_rows],\
        [chr(ord('B') + delta), 2, chr(ord('B') + delta), data_rows], \
        [chr(ord('C') + delta), 2, chr(ord('F') + delta), data_rows],\
        [label['BD-PSNR'][1], 2, label['BD-PSNR'][2], data_rows], \
        [label['SPEED-UP'][1], 2, label['SPEED-UP'][2], data_rows]])

    if type == 'all':
        set_solid_border(sheet,[[label['BD-SSIM'][1], 1, label['BD-SSIM'][2], 1], \
            [label['BD-VMAF'][1], 1, label['BD-VMAF'][2], 1]])
        set_solid_border(sheet,[['G', 2, 'J', data_rows], \
            [chr(ord('G') + delta), 2, chr(ord('J') + delta), data_rows], \
            ['K', 2, 'K', data_rows], [chr(ord('K') + delta), 2, chr(ord('K') + delta), data_rows], \
            ['L', 2, 'L', data_rows], [chr(ord('L') + delta), 2, chr(ord('L') + delta), data_rows]])
        set_solid_border(sheet,[[label['BD-SSIM'][1], 2, label['BD-SSIM'][2], data_rows], \
            [label['BD-VMAF'][1], 2, label['BD-VMAF'][2], data_rows]])
    elif type == 'psnr':
        set_solid_border(sheet,[['G', 2, 'G', data_rows], [chr(ord('G') + delta), 2, chr(ord('G') + delta), data_rows]])

    #fill green colors and change data type
    green_fill = PatternFill("solid", start_color='CCFFFF')
    for row in sheet.iter_rows(min_row=3, max_row=(data_rows), min_col=3, max_col=(label['test'][-1]-2)):
        for cell in row:
            cell.fill = green_fill
            cell.data_type = 'n'
    for row in sheet.iter_rows(min_row=3, max_row=(data_rows), min_col=3+delta, max_col=(label['BD-PSNR'][-1]-2)):
        for cell in row:
            cell.fill = green_fill
            cell.data_type = 'n'

def csv_to_xlsx(csv_file, excel_file, sheet_name, type = 'psnr', anchor='ANCHOR', test='TEST', recreate=False):
    print("--------------------START csv_to_xlsx----------------------------------")
    print("csv_file:" + csv_file)
    print("excel_file:" + excel_file)
    print("sheet_name:" + sheet_name)
    print("type:" + type)
    print("anchor_name:" + anchor)
    print("test_name:" + test)
    with open(csv_file, 'rb') as f:
        read = csv.reader(f)
        if not os.path.exists(excel_file) or recreate == True:
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.create_sheet(sheet_name)
        label = gen_lable_index(type)
        draw_header(sheet, label, anchor, test, type)
        draw_second_row(sheet, label, anchor, test, type)
        csvFile = open(csv_file, "r")
        reader = csv.reader(csvFile)
        endrow = 1
        entries = 1
        for i,line in enumerate(reader):
            for c in range(len(line)):
                sheet.cell(row=i+2, column=c+1).value = line[c]
            if sheet.cell(row=i+2, column=1).value != "Average":
                entries = entries + 1
            endrow = endrow + 1
        
        csvFile.close()

        # format data and draw borders
        format_excel(sheet, label, entries, endrow, type)
        workbook.save(excel_file)
    print("------------------------FINISHED---------------------------------------")
    return endrow

def gen_summary_sheet(excel_file, type = "psnr"):
    if not os.path.exists(excel_file):
        print("ERROR: excel_file not exists")
        return
    else:
        workbook = openpyxl.load_workbook(excel_file)
    title_font = Font(name='Times New Roman', bold = True)
    normal_font = Font(name='Arial', bold = False)
    align = Alignment(horizontal='center', vertical='center')
    label = gen_lable_index(type)
    end_label,delta = get_delta_endlabel(type)
    sheet = workbook['Sheet']
    Summary = "Summary"
    sheet.title = "Summary"
    start_row = 2
    for iter_sheet in workbook:
        if (iter_sheet.title == Summary):
            for col in range(2, delta+2):
                sheet.column_dimensions[get_column_letter(col)].width = 15.0
            continue
        #sub table title
        sheet.merge_cells("C" + str(start_row) + ":" + get_column_letter(delta+1) + str(start_row))
        sheet.cell(row=start_row, column=3).value = iter_sheet.title.upper()
        sheet.cell(row=start_row, column=3).font = title_font
        sheet.cell(row=start_row, column=3).alignment = align
        sheet.cell(row=start_row, column=3).fill = PatternFill("solid", start_color='E6B8B7')
        #version
        sheet.merge_cells("C" + str(start_row+1) + ":" + get_column_letter(delta+1) + str(start_row+1))
        anchor_name = 'ANCHOR[' + iter_sheet.cell(row=1, column=label['anchor'][3]).value + ']'
        test_name = 'TESTED[' + iter_sheet.cell(row=1, column=label['test'][3]).value + ']'
        sheet.cell(row=start_row+1, column=3).value = anchor_name + "    " + test_name
        sheet.cell(row=start_row+1, column=3).font = title_font
        sheet.cell(row=start_row+1, column=3).fill = PatternFill("solid", start_color='E6B8B7')
        sheet.cell(row=start_row+1, column=3).alignment = align
        #PSNR-title
        sheet.cell(row=start_row+2, column=3).value = "CLASS"
        sheet.cell(row=start_row+2, column=3).font = title_font
        sheet.cell(row=start_row+2, column=3).alignment = align
        sheet.cell(row=start_row+2, column=4).value = "Y-BD(PSNR)"
        sheet.cell(row=start_row+2, column=4).font = title_font
        sheet.cell(row=start_row+2, column=4).alignment = align
        sheet.cell(row=start_row+2, column=5).value = "U-BD(PSNR)"
        sheet.cell(row=start_row+2, column=5).font = title_font
        sheet.cell(row=start_row+2, column=5).alignment = align
        sheet.cell(row=start_row+2, column=6).value = "V-BD(PSNR)"
        sheet.cell(row=start_row+2, column=6).font = title_font
        sheet.cell(row=start_row+2, column=6).alignment = align
        sheet.cell(row=start_row+2, column=7).value = "C-BD(PSNR)"
        sheet.cell(row=start_row+2, column=7).font = title_font
        sheet.cell(row=start_row+2, column=7).alignment = align

        #SPEED-UP-title
        sheet.cell(row=start_row+2, column=delta+1).value = "SPEED-UP"
        sheet.cell(row=start_row+2, column=delta+1).font = title_font
        sheet.cell(row=start_row+2, column=delta+1).alignment = align

        #BD-RATE-DATA
        bd_col = label['BD-PSNR'][3]
        max_row = iter_sheet.max_row
        ave_begin_row = max_row
        for i_row in range(max_row, 0, -1):
            if iter_sheet.cell(row=i_row, column=1).value == "Average":
                ave_begin_row = ave_begin_row - 1
            else:
                break
        
        row_offset = 3
        class_rows = max_row - ave_begin_row 
        for i_row in range(ave_begin_row+1, max_row+1, 1):
            change_row = start_row + row_offset
            sheet.cell(row=change_row, column=3).value = iter_sheet.cell(row=i_row, column=2).value
            sheet.cell(row=change_row, column=3).font = title_font
            sheet.cell(row=change_row, column=3).alignment = align
            sheet.cell(row=change_row, column=4).value = iter_sheet.cell(row=i_row, column=bd_col).value
            sheet.cell(row=change_row, column=4).font = normal_font
            sheet.cell(row=change_row, column=4).alignment = align
            sheet.cell(row=change_row, column=5).value = iter_sheet.cell(row=i_row, column=bd_col+1).value
            sheet.cell(row=change_row, column=5).font = normal_font
            sheet.cell(row=change_row, column=5).alignment = align
            sheet.cell(row=change_row, column=6).value = iter_sheet.cell(row=i_row, column=bd_col+2).value
            sheet.cell(row=change_row, column=6).font = normal_font
            sheet.cell(row=change_row, column=6).alignment = align
            sheet.cell(row=change_row, column=7).value = iter_sheet.cell(row=i_row, column=bd_col+3).value
            sheet.cell(row=change_row, column=7).font = normal_font
            sheet.cell(row=change_row, column=7).alignment = align
            #SPEED-UP
            sheet.cell(row=change_row, column=delta+1).value = iter_sheet.cell(row=i_row, column=label["SPEED-UP"][3]).value
            sheet.cell(row=change_row, column=delta+1).font = normal_font
            sheet.cell(row=change_row, column=delta+1).alignment = align
            row_offset = row_offset + 1



        if (type == "all"):
            #PSNR-title
            sheet.cell(row=start_row+2, column=8).value = "Y-BD(SSIM)"
            sheet.cell(row=start_row+2, column=8).font = title_font
            sheet.cell(row=start_row+2, column=8).alignment = align
            sheet.cell(row=start_row+2, column=9).value = "U-BD(SSIM)"
            sheet.cell(row=start_row+2, column=9).font = title_font
            sheet.cell(row=start_row+2, column=9).alignment = align
            sheet.cell(row=start_row+2, column=10).value = "V-BD(SSIM)"
            sheet.cell(row=start_row+2, column=10).font = title_font
            sheet.cell(row=start_row+2, column=10).alignment = align
            sheet.cell(row=start_row+2, column=11).value = "C-BD(SSIM)"
            sheet.cell(row=start_row+2, column=11).font = title_font
            sheet.cell(row=start_row+2, column=11).alignment = align
            sheet.cell(row=start_row+2, column=12).value = "BD(VMAF)"
            sheet.cell(row=start_row+2, column=12).font = title_font
            sheet.cell(row=start_row+2, column=12).alignment = align
            #BD-RATE-DATA(SSIM)
            #max_row = iter_sheet.max_row
            row_offset = 3
            for i_row in range(ave_begin_row+1, max_row+1, 1):
                bd_col = label['BD-SSIM'][3]
                change_row = start_row + row_offset
                sheet.cell(row=change_row, column=8).value = iter_sheet.cell(row=i_row, column=bd_col).value
                sheet.cell(row=change_row, column=8).font = normal_font
                sheet.cell(row=change_row, column=8).alignment = align
                sheet.cell(row=change_row, column=9).value = iter_sheet.cell(row=i_row, column=bd_col+1).value
                sheet.cell(row=change_row, column=9).font = normal_font
                sheet.cell(row=change_row, column=9).alignment = align
                sheet.cell(row=change_row, column=10).value = iter_sheet.cell(row=i_row, column=bd_col+2).value
                sheet.cell(row=change_row, column=10).font = normal_font
                sheet.cell(row=change_row, column=10).alignment = align
                sheet.cell(row=change_row, column=11).value = iter_sheet.cell(row=i_row, column=bd_col+3).value
                sheet.cell(row=change_row, column=11).font = normal_font
                sheet.cell(row=change_row, column=11).alignment = align
                #BD-RATE-DATA(VMAF)
                bd_col = label['BD-VMAF'][3]
                sheet.cell(row=change_row, column=12).value = iter_sheet.cell(row=i_row, column=bd_col).value
                sheet.cell(row=change_row, column=12).font = normal_font
                sheet.cell(row=change_row, column=12).alignment = align
                row_offset = row_offset + 1
            set_solid_border(sheet, [["H", start_row+2, "K", start_row+2+class_rows], \
                ["L", start_row+2, "L", start_row+2+class_rows]])

        set_solid_border(sheet, [["C", start_row, get_column_letter(delta+1), start_row+1], \
            ["C", start_row+2, get_column_letter(delta+1), start_row+2], \
            ["C", start_row+2, get_column_letter(delta+1), start_row+1+class_rows], \
            ["C", start_row+2+class_rows, get_column_letter(delta+1), start_row+2+class_rows], \
            ["C", start_row+2, "D", start_row+2+class_rows], \
            ["D", start_row+2, "G", start_row+2+class_rows], \
            [get_column_letter(delta+1), start_row+2, get_column_letter(delta+1), start_row+2+class_rows]])
        start_row += (class_rows + 6)

    workbook.save(excel_file)

if __name__ == '__main__':
    csv_to_xlsx(sys.argv[1], sys.argv[2], "camera", "psnr")